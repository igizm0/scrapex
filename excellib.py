import sys, os, threading
from xlwt import *
import xlrd

try:
	import openpyxl
		
except:
	raise	

def savexlsx(filepath, data):
	book = openpyxl.Workbook()
	sheet = book.active
	sheet.title = 'Sheet1'
		
	rowindex = 0
	for i, r in enumerate (data):
		heahers = []
		values = []
		for j, col in enumerate(r):
			if j % 2 == 0:
				heahers.append(col)
			else:			
				if isinstance(col, basestring):
					col = col.strip()			
				values.append(col)
		if i==0:
			#write headers
			for colindex, h in enumerate(heahers):
				sheet.cell(row=0,column=colindex).value = h

		rowindex += 1
		for colindex, value in enumerate(values):
			sheet.cell(row=rowindex,column=colindex).value = value if value is not None else ''

	book.save(filepath)		

def savexls(filepath, data):
	
	book = Workbook(encoding="utf-8")
	
	sheet = book.add_sheet("Sheet1")
	style = XFStyle()
	style.num_format_str = '0.00'
	rowindex = 0
	for i, r in enumerate (data):
		heahers = []
		values = []
		for j, col in enumerate(r):
			if j % 2 == 0:
				heahers.append(col)
			else:			
				if isinstance(col, basestring):
					col = col.strip()			
				values.append(col)
		if i==0:
			#write headers
			for colindex, h in enumerate(heahers):
				sheet.write(0,colindex,h)

		rowindex += 1
		for colindex, value in enumerate(values):
			if value is not None:			
				sheet.write(rowindex,colindex,value, style)	
			else:				
				sheet.write(rowindex,colindex,'', style)	

	book.save(filepath)			

def csvdatatoxls(filepath, data):
	
	book = Workbook(encoding="utf-8")
	
	sheet = book.add_sheet("sheet1")
	style = XFStyle()
	style.num_format_str = '0.00'
	rowindex = -1
	for r in data:
		rowindex += 1
		
		for colindex, value in enumerate(r):
			sheet.write(rowindex,colindex,value, style)	

	book.save(filepath)			

def readxlsxsheet(filepath, index=0):
	wb = openpyxl.load_workbook(filepath)
	
	sheet = wb.worksheets[index]	
	data = []
	for row in sheet.rows:
		r = []
		for cell in row:
			r.append(cell.value)
		data.append(r)	

	return data	

def readsheet(filepath, restype='list', index=0):
	"""
	restype: list, dict
	"""
	data = []
	if filepath.lower().endswith('.xlsx'):
		data = readxlsxsheet(filepath)
	else:	
		book = xlrd.open_workbook(filepath)	
		sheet1 = book.sheet_by_index(index)

		for i in range(sheet1.nrows):
			r = sheet1.row_values(i)		
			data.append(r)
	
	fields =[]

	for field in data[0]:
		if not field: break #every fields after the first None field are ignored
		fields.append(field)

	
	if restype == 'list':
		if len(fields)<len(data[0]):
			cleaneddata = [r[0:len(fields)] for r in data]
			return cleaneddata
		else:	
			return data

	
	rs = []
	rowindex = 1
	for r in data[1:]:
		rowindex += 1
		if len(r) != len(data[0]):
			raise Exception("Inconsistent row length at row#: %s" % rowindex)

		row = {}	
		for i, value in enumerate(r):
			if i== len(fields): break			
			row.update({fields[i]: value})
		rs.append(row)	
		
	return rs	


if __name__ == '__main__':
	print 'test'
	data = readsheet('d:/scrape/inv/testingdata/7.20.14_foll_wholesale.xlsx', 'list')
	print data[0]




